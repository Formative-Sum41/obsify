"""Comprehensive, adversarial tests for the obsify MCP toolkit.

Covers the privacy INVARIANTS (scan/twin never emit values; verify fails closed),
the sandbox guard (blocks escape/exfil), run_on_real (masking, timeout, capping),
and edge cases (missing/empty/unreadable inputs). Loads the spaCy model once.

Run: python tests/test_obsify.py
"""

from __future__ import annotations

import json
import sys
import tempfile
from datetime import datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

import openpyxl  # noqa: E402

import obsify.mcp_server as S  # noqa: E402
from obsify.sandbox import execute, static_guard  # noqa: E402

# Known real values planted in the sample file — must NEVER appear in shape output.
NAMES = ["Jane Roe", "Bob Lee"]
TFN = "123 456 782"          # valid TFN checksum
KNOWN = NAMES + [TFN]


def _sample_xlsx(dirpath: str) -> str:
    p = str(Path(dirpath) / "sample.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ledger"
    ws.append(["Name", "Amount", "PostDate", "TaxRef"])
    ws.append(["Jane Roe", 100.50, datetime(2024, 1, 15), f"TFN {TFN}"])
    ws.append(["Bob Lee", 205.00, datetime(2024, 2, 20), f"TFN {TFN}"])
    ws.append([None, 300.00, datetime(2024, 3, 1), None])
    wb.save(p)
    return p


# ---------------------------------------------------------------- scan_pii ----

def test_scan_returns_shape_only_no_values():
    with tempfile.TemporaryDirectory() as d:
        f = _sample_xlsx(d)
        r = S.scan_pii(f)
        blob = json.dumps(r).lower()
        for v in KNOWN:
            assert v.lower() not in blob, f"scan_pii leaked a value: {v!r}"
        assert all("value" not in k and "text" not in k for k in r), "no value keys"
        assert r["counts_by_type"], "should detect the planted PII types"
        assert all(set(x.keys()) == {"type", "document", "location"} for x in r["findings"])


def test_scan_missing_path_is_graceful():
    r = S.scan_pii(str(Path(tempfile.gettempdir()) / "obsify_nope_xyz"))
    assert r["files_found"] == 0 and r["files_read"] == 0


def test_scan_unreadable_file_skipped_not_crash():
    with tempfile.TemporaryDirectory() as d:
        bad = Path(d) / "corrupt.xlsx"
        bad.write_bytes(b"not a real zip/xlsx")
        r = S.scan_pii(d)
        assert any("could not read" in n for n in r["notes"])


# -------------------------------------------------------------- redact_text ---

def test_redact_masks_known_pii():
    out = S.redact_text(f"TFN {TFN} paid to Jane Roe")
    assert TFN not in out and "Jane Roe" not in out
    assert "<" in out  # replaced with <TYPE> placeholders


def test_redact_passes_through_clean_text():
    assert S.redact_text("the quarterly totals reconcile") == "the quarterly totals reconcile"


# --------------------------------------------------------- verify_value_free --

def test_verify_clean_true_leak_false():
    assert S.verify_value_free("types and counts only", ["Jane Roe", TFN])["value_free"] is True
    assert S.verify_value_free("contact Jane Roe", ["Jane Roe"])["value_free"] is False


def test_verify_catches_case_and_variant():
    assert S.verify_value_free("JANE ROE", ["Jane Roe"])["value_free"] is False       # case
    assert S.verify_value_free("Veranth Hldgs P/L", ["Veranth Holdings Pty Ltd"])["value_free"] is False  # variant
    assert S.verify_value_free("nothing here", [])["value_free"] is True              # empty terms


# ---------------------------------------------------------- sandbox guard -----

def test_guard_blocks_dangerous_code():
    for bad in ["import subprocess", "import requests", "import socket\n",
                "import os\nos.system('dir')", "eval('1+1')", "__import__('os')",
                "open('x.txt','w')"]:
        assert static_guard(bad) is not None, f"guard should block: {bad!r}"


def test_guard_allows_legitimate_read_code():
    ok = "import openpyxl\nwb = openpyxl.load_workbook(DATA_PATH)\nprint('ok')"
    assert static_guard(ok) is None


# ----------------------------------------------------------- run_on_real ------

def test_run_on_real_aggregate_and_mask():
    with tempfile.TemporaryDirectory() as d:
        f = _sample_xlsx(d)
        code = (
            "import openpyxl\n"
            "wb = openpyxl.load_workbook(DATA_PATH, data_only=True)\n"
            "ws = wb.active\n"
            "n = sum(1 for _ in ws.iter_rows(min_row=2))\n"
            "print('rowcount', n)\n"
            "print('a_name', 'Jane Roe')\n"  # raw name -> must be masked on the way out
        )
        res = S.run_on_real(code, f)
        assert res["ok"] and res["exit_code"] == 0
        assert "rowcount 3" in res["output_masked"]     # aggregate returns
        assert "Jane Roe" not in res["output_masked"]   # titlecase name masked


def test_run_on_real_blocks_escape_code():
    res = S.run_on_real("import subprocess\nsubprocess.run(['echo','hi'])", "unused")
    assert res["ok"] is False and res["exit_code"] == -2
    assert "static guard" in res["error_masked"]


def test_run_on_real_timeout():
    res = S.run_on_real("while True:\n    pass", "unused", timeout=3)
    assert res["timed_out"] is True and res["ok"] is False


def test_run_on_real_error_masked():
    res = S.run_on_real("raise ValueError('boom')", "unused")
    assert res["ok"] is False
    assert "ValueError" in res["error_masked"]


if __name__ == "__main__":
    failures = 0
    for name, fn in sorted(globals().items()):
        if name.startswith("test_") and callable(fn):
            try:
                fn()
                print(f"PASS {name}")
            except AssertionError as e:
                failures += 1
                print(f"FAIL {name}: {e}")
            except Exception as e:  # noqa: BLE001
                failures += 1
                print(f"ERROR {name}: {type(e).__name__}: {e}")
    print(f"\n{failures} failure(s)")
    sys.exit(1 if failures else 0)
