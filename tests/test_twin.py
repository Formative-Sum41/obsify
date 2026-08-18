"""Tests for the compute-to-data pieces: synthetic twin (obsify/twin.py) and the
execution sandbox (obsify/sandbox.py).

Run: python tests/test_twin.py
"""

from __future__ import annotations

import sys
import tempfile
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

import openpyxl  # noqa: E402

from obsify.sandbox import execute  # noqa: E402
from obsify.twin import make_synthetic_twin  # noqa: E402

# Distinctive planted values (all synthetic) that a faithful twin must NOT copy —
# the test generates its own source workbook, so the suite needs no committed binary.
SEED_VALUES = ["zzcalderwell", "priya qnair", "zbluehaven ltd", "oskestra pty", "0433 221 190"]


def _seed_xlsx(dirpath: str) -> str:
    """A small schema-realistic GL-style workbook planted with SEED_VALUES."""
    p = str(Path(dirpath) / "seed.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Journals"
    ws.append(["Entity", "Contact", "Counterparty", "Phone", "Amount"])
    ws.append(["zzcalderwell", "priya qnair", "zbluehaven ltd", "0433 221 190", 100.50])
    ws.append(["zzcalderwell", "priya qnair", "oskestra pty", "0433 221 190", 205.00])
    for i in range(3, 30):
        ws.append([f"zzcalderwell", "priya qnair", "oskestra pty", "0433 221 190", 10.0 * i])
    wb.save(p)
    return p


def test_twin_preserves_schema():
    with tempfile.TemporaryDirectory() as d:
        out = str(Path(d) / "twin.xlsx")
        schema = make_synthetic_twin(_seed_xlsx(d), out, cap_rows=50)
        assert Path(out).exists()
        assert schema["sheets"], "twin should describe at least one sheet"
        cols = schema["sheets"][0]["columns"]
        assert any(c["name"] for c in cols), "column headers (schema) preserved"


def test_twin_leaks_no_real_values():
    with tempfile.TemporaryDirectory() as d:
        out = str(Path(d) / "twin.xlsx")
        make_synthetic_twin(_seed_xlsx(d), out, cap_rows=200)
        wb = openpyxl.load_workbook(out)
        blob = " ".join(str(c.value) for ws in wb for row in ws.iter_rows()
                        for c in row if c.value is not None).lower()
        for v in SEED_VALUES:
            assert v not in blob, f"planted value leaked into twin: {v!r}"


def test_sandbox_runs_and_captures_output():
    res = execute("print('rows =', 40 + 2)", data_path="unused")
    assert res.exit_code == 0
    assert "rows = 42" in res.stdout


def test_sandbox_static_guard_blocks_network():
    res = execute("import socket\nsocket.create_connection(('example.com', 80))",
                  data_path="unused")
    assert res.exit_code == -2
    assert "static guard" in res.stderr and "socket" in res.stderr
    assert "CONNECTED" not in res.stdout


def test_sandbox_injects_data_path():
    res = execute("print('path is', DATA_PATH)", data_path="C:/some/real/file.xlsx")
    assert "C:/some/real/file.xlsx" in res.stdout


if __name__ == "__main__":
    failures = 0
    for name, fn in sorted(globals().items()):
        if name.startswith("test_") and callable(fn):
            try:
                fn()
                print(f"PASS {name}")
            except AssertionError as e:
                failures += 1
                print(f"FAIL {name}: {e}")
    print(f"\n{failures} failure(s)")
    sys.exit(1 if failures else 0)
