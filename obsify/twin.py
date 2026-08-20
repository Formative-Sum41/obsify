"""Synthetic twin generation for compute-to-data.

Produce a FAITHFUL FAKE surrogate of a real Excel workbook: the schema (sheet
names, column headers, column count, row counts) and column *types/ranges/
cardinality* are preserved, but every DATA value is freshly generated — no real
value is copied. A frontier model reasons against the twin (writes code that
references columns, handles types and edge cases); that code then runs on the
REAL data via the sandbox, so substance never enters the model's context.

Scope: Excel (.xlsx/.xlsm) — the structured-data case. Values are faked by
inferred type (id / number / date / text), preserving order-of-magnitude and
date ranges so aggregates on the twin are plausible but meaningless. Row counts
are reported truthfully but the twin body is capped for size.
"""

from __future__ import annotations

import random
from dataclasses import dataclass
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl


@dataclass
class ColumnProfile:
    header: str
    kind: str          # id | number | date | text | empty
    nulls: float       # fraction empty
    minimum: float | None = None
    maximum: float | None = None
    decimals: int = 0
    distinct: int = 0
    date_min: datetime | None = None
    date_max: datetime | None = None


def _detect_header_row(rows: list[list], scan: int = 15) -> int:
    """Header = the row (within the first `scan`) with the most non-empty STRING
    cells — robust to title/metadata rows above the real header."""
    best_i, best_strs = 0, -1
    for i, row in enumerate(rows[:scan]):
        strs = sum(1 for v in row if isinstance(v, str) and v.strip())
        if strs > best_strs:
            best_i, best_strs = i, strs
    return best_i


def _profile_column(values: list) -> ColumnProfile:
    vals = [v for v in values if v is not None and v != ""]
    nulls = 1.0 - (len(vals) / len(values)) if values else 1.0
    if not vals:
        return ColumnProfile(header="", kind="empty", nulls=nulls)
    if all(isinstance(v, datetime) for v in vals):
        return ColumnProfile(header="", kind="date", nulls=nulls,
                             date_min=min(vals), date_max=max(vals))
    nums = [v for v in vals if isinstance(v, (int, float)) and not isinstance(v, bool)]
    if len(nums) == len(vals):
        ints = [int(v) for v in nums if float(v).is_integer()]
        monotonic = len(ints) == len(nums) and all(
            ints[i] <= ints[i + 1] for i in range(len(ints) - 1)) and len(ints) > 2
        decimals = max((len(str(v).split(".")[1]) for v in nums
                        if isinstance(v, float) and "." in str(v)), default=0)
        return ColumnProfile(header="", kind="id" if monotonic else "number",
                             nulls=nulls, minimum=min(nums), maximum=max(nums),
                             decimals=min(decimals, 4))
    strs = [str(v) for v in vals]
    # semantic text subtype: emails get type-faithful fakes so twin columns are usable
    # for developing parsing/validation code (its stated purpose), not just shape.
    if sum(("@" in s and "." in s.rsplit("@", 1)[-1]) for s in strs) >= max(1, len(strs) // 2):
        return ColumnProfile(header="", kind="email", nulls=nulls)
    return ColumnProfile(header="", kind="text", nulls=nulls, distinct=len(set(strs)))


def _fake(profile: ColumnProfile, i: int, rng: random.Random):
    if profile.nulls and rng.random() < profile.nulls:
        return None
    k = profile.kind
    if k == "id":
        return 100000 + i
    if k == "number":
        lo, hi = profile.minimum or 0, profile.maximum or 1
        v = rng.uniform(lo, hi)
        return round(v, profile.decimals) if profile.decimals else int(v)
    if k == "date":
        lo = profile.date_min or datetime(2020, 1, 1)
        hi = profile.date_max or datetime(2024, 1, 1)
        span = max((hi - lo).days, 1)
        return lo + timedelta(days=rng.randint(0, span))
    if k == "email":
        return f"user{100000 + i}@example.com"
    if k == "text":
        return f"val_{i}_{rng.randint(1, max(profile.distinct, 1))}"
    return None


def make_synthetic_twin(path: str, out: str, cap_rows: int = 2000, seed: int = 7) -> dict:
    """Write a synthetic twin of the Excel workbook at `path` to `out`. Returns a
    schema summary (sheets, columns, inferred types, TRUE row counts). No real
    value is copied. Column headers are preserved (schema); if a header itself is
    sensitive that is the caller's responsibility to review (headers are usually
    generic field names)."""
    rng = random.Random(seed)
    src_wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    out_wb = openpyxl.Workbook()
    out_wb.remove(out_wb.active)
    schema = []
    for ws in src_wb.worksheets:
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        if not rows:
            out_wb.create_sheet(ws.title)
            schema.append({"sheet": ws.title, "rows": 0, "columns": []})
            continue
        h = _detect_header_row(rows)
        header = rows[h]
        data_rows = rows[h + 1:]
        ncols = max((len(r) for r in rows), default=0)
        profiles: list[ColumnProfile] = []
        for c in range(ncols):
            col_vals = [r[c] if c < len(r) else None for r in data_rows]
            p = _profile_column(col_vals)
            p.header = str(header[c]) if c < len(header) and header[c] is not None else f"col_{c}"
            profiles.append(p)
        # write twin sheet: header + capped fake data
        out_ws = out_wb.create_sheet(ws.title)
        out_ws.append([p.header for p in profiles])
        n = min(len(data_rows), cap_rows)
        for i in range(n):
            out_ws.append([_fake(p, i, rng) for p in profiles])
        schema.append({
            "sheet": ws.title,
            "rows": len(data_rows),          # TRUE row count
            "twin_rows": n,                  # rows written (capped)
            "columns": [{"name": p.header, "type": p.kind, "nulls": round(p.nulls, 2)}
                        for p in profiles],
        })
    src_wb.close()
    Path(out).parent.mkdir(parents=True, exist_ok=True)
    out_wb.save(out)
    return {"twin_path": out, "sheets": schema}
